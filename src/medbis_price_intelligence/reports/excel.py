from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from medbis_price_intelligence.reports.intelligence import (
    MatchDataService,
    PriceIntelligenceService,
    ProductPriceInsight,
)


class ExcelReportGenerator:
    """Creates MedBIS price intelligence Excel workbooks."""

    SHEET_ORDER = [
        "Summary",
        "Competitor Prices",
        "Products Above Market",
        "Margin Opportunities",
        "Historical Prices",
        "Products Without Matches",
    ]

    def __init__(self, session: Session) -> None:
        self.session = session

    def generate(self, output_path: Path) -> Path:
        """Generate the complete report workbook."""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        insights = PriceIntelligenceService(self.session).calculate()
        match_data = MatchDataService(self.session)

        sheets = {
            "Summary": self._summary(insights),
            "Competitor Prices": pd.DataFrame(match_data.competitor_price_rows()),
            "Products Above Market": self._filter_insights(
                insights,
                {"above_competitor", "above_market_average"},
            ),
            "Margin Opportunities": self._filter_insights(
                insights,
                {"medbis_cheapest", "significantly_below_market"},
                require_opportunity=True,
            ),
            "Historical Prices": pd.DataFrame(match_data.historical_price_rows()),
            "Products Without Matches": self._filter_insights(insights, {"no_competitor_match"}),
        }

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name in self.SHEET_ORDER:
                sheets[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
            self._format_workbook(writer.book)

        return output_path

    @staticmethod
    def _summary(insights: list[ProductPriceInsight]) -> pd.DataFrame:
        counts = {
            "Total Products": len(insights),
            "With Competitor Matches": sum(1 for row in insights if row.competitor_count > 0),
            "MedBIS Cheapest": sum(1 for row in insights if row.market_position == "medbis_cheapest"),
            "Above Competitor": sum(1 for row in insights if row.market_position == "above_competitor"),
            "Margin Opportunities": sum(1 for row in insights if row.margin_opportunity),
            "Without Matches": sum(1 for row in insights if row.market_position == "no_competitor_match"),
        }
        return pd.DataFrame([{"Metric": key, "Value": value} for key, value in counts.items()])

    def _filter_insights(
        self,
        insights: list[ProductPriceInsight],
        positions: set[str],
        require_opportunity: bool = False,
    ) -> pd.DataFrame:
        rows = [
            insight
            for insight in insights
            if insight.market_position in positions
            and (not require_opportunity or insight.margin_opportunity is not None)
        ]
        return pd.DataFrame([self._insight_to_dict(row) for row in rows])

    @staticmethod
    def _insight_to_dict(insight: ProductPriceInsight) -> dict[str, object]:
        return {
            "SKU": insight.sku,
            "Product": insight.product_name,
            "Brand": insight.brand,
            "MedBIS Price": insight.medbis_price,
            "Lowest Competitor Price": insight.lowest_competitor_price,
            "Average Competitor Price": insight.average_competitor_price,
            "Competitor Count": insight.competitor_count,
            "Market Position": insight.market_position,
            "Suggested Price": insight.suggested_price,
            "Margin Opportunity": insight.margin_opportunity,
        }

    @staticmethod
    def _format_workbook(workbook: object) -> None:
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            for column in worksheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = min(
                    max(max_length + 2, 12),
                    55,
                )

